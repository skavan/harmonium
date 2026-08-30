#!/usr/bin/env python3
"""Regenerate <version>/astrion-remote-map.md + <version>/"KeyCodes Astrion.xlsx"
FROM each version's key_mapper.zip - the KeyMapper backup is the truth, these
docs are its rendering (they rotted twice when maintained by hand; v0.83.11).

Version-aware and self-extracting: it reads data.json straight out of
key_mapper.zip (writing a readable data.json copy beside it), so there is no
manual extract step.

  python gen-map-docs.py            # regenerate every version (v1, v2, ...)
  python gen-map-docs.py v2         # just that version's folder
  python gen-map-docs.py <path>     # an explicit version folder

xlsx needs openpyxl (pip install openpyxl); without it the markdown still
regenerates. pull-keymapper offers to run this after a pull."""
import base64
import json
import sys
import zipfile
from pathlib import Path

HERE = Path(__file__).parent  # remotes/astrion/keymapper

# the vocabulary this remote actually uses - extend as keys appear.
# labels are the PHYSICAL buttons ("what is F2?" - Power), per
# remotes/astrion/facts.md. F-number kept in parentheses.
KEYCODE = {  # Android input keycode -> (constant, physical label)
    4: ("KEYCODE_BACK", "Back"), 24: ("KEYCODE_VOLUME_UP", "Volume Up"),
    25: ("KEYCODE_VOLUME_DOWN", "Volume Down"), 82: ("KEYCODE_MENU", "Menu"),
    131: ("KEYCODE_F1", "Home (F1)"), 132: ("KEYCODE_F2", "Power (F2)"),
    134: ("KEYCODE_F4", "Lightbulb·REW (F4)"),
    135: ("KEYCODE_F5", "Curtains·Play/Pause (F5)"),
    136: ("KEYCODE_F6", "Music·Stop (F6)"),
    137: ("KEYCODE_F7", "Climate·FWD (F7)"),
    138: ("KEYCODE_F8", "Red (F8)"), 139: ("KEYCODE_F9", "Green (F9)"),
    140: ("KEYCODE_F10", "Blue (F10)"), 141: ("KEYCODE_F11", "Yellow (F11)"),
    164: ("KEYCODE_VOLUME_MUTE", "Volume Mute"),
    92: ("KEYCODE_PAGE_UP", "Channel Up"),
    93: ("KEYCODE_PAGE_DOWN", "Channel Down"),
    21: ("KEYCODE_DPAD_LEFT", "D-pad Left"),
    22: ("KEYCODE_DPAD_RIGHT", "D-pad Right"),
}
OUTKEY = {  # KEY_EVENT output keycode -> (constant, printed character)
    68: ("KEYCODE_GRAVE", "`"), 69: ("KEYCODE_MINUS", "-"),
    70: ("KEYCODE_EQUALS", "="), 71: ("KEYCODE_LEFT_BRACKET", "["),
    72: ("KEYCODE_RIGHT_BRACKET", "]"), 74: ("KEYCODE_SEMICOLON", ";"),
    18: ("KEYCODE_POUND", "#"), 81: ("KEYCODE_PLUS", "+"),
    75: ("KEYCODE_APOSTROPHE", "'"), 76: ("KEYCODE_SLASH", "/"),
    55: ("KEYCODE_COMMA", ","), 56: ("KEYCODE_PERIOD", "."),
    77: ("KEYCODE_AT", "@"), 86: ("KEYCODE_MEDIA_STOP", "MediaStop"),
    142: ("KEYCODE_F12", "F12"),
}
APP = {
    "io.github.sds100.keymapper": "Open Key Mapper",
    "de.ozerov.fully": "Open Fully Kiosk Browser",
    "fr.neamar.kiss": "Open KISS Launcher",
    "com.mediatek.filemanager": "Open File Manager",
    "com.aiks.HaRemote": "Open HaRemote",
    "com.android.browser": "Open Android Browser",
    "io.homeassistant.companion.android.minimal": "Open Home Assistant Minimal",
}
SYS = {"go_back": "Android Go back", "go_home": "Android Go home"}
CLICK = {0: "press", 1: "long-press", 2: "double-press"}

# The four coloured keys, in remote layout order. The "what they do"
# Notes line is DERIVED from the mappings below (it used to be hand-typed
# and silently disagreed with the table for v2 - v0.86).
COLOUR = [(138, "Red"), (139, "Green"), (140, "Blue"), (141, "Yellow")]

RAW = [
    ("D-pad ▲ ▼ ◀ ▶", "Arrow keys", "move panel focus · drive the device on TV pages"),
    ("OK (center)", "Enter", "activate the focused card · hold = grab the D-pad"),
    ("Home (F1) tap", "`F1`", "Harmonium home — up one level"),
    ("Power (F2) tap", "`F2`", "power — end/start the page's activity (confirm)"),
    ("Channel ▲/▼ tap", "PageUp / PageDown", "jump sections · on TV pages: borrow the D-pad for the panel"),
    ("Lightbulb·REW (F4)", "`F4`", "astrion profile: Lights shortcut · v2 recipe: ⏮ previous"),
    ("Curtains·Play/Pause (F5)", "`F5`", "astrion: Covers shortcut · v2: ⏯"),
    ("Music·Stop (F6)", "`F6`", "astrion: Music shortcut · v2: ⏹"),
    ("Climate·FWD (F7)", "`F7`", "astrion: Climate shortcut · v2: ⏭ next"),
]


def extras_of(a):
    return {x.get("id"): x.get("data") for x in a.get("extras", [])}


def decode_shell(a):
    """The exact shell command a SHELL_COMMAND action runs. Key Mapper stores
    it base64-encoded (with embedded newlines, which b64decode ignores)."""
    try:
        return base64.b64decode(a.get("data", "")).decode("utf-8", "replace").strip()
    except Exception:
        return "(could not decode)"


def describe_action_cell(a):
    """(output_text, output_keycode/package, output_constant) for the table.
    Human-readable - shell commands and sounds use the descriptions Key Mapper
    stored in the action's extras (falling back to a decoded command / the uid),
    so nothing shows as raw base64. The full decoded shell command lives in its
    own section (see shell_section)."""
    t = a["type"]
    ex = extras_of(a)
    if t == "APP":
        return (APP.get(a["data"], "Open " + a["data"]), a["data"], "—")
    if t == "KEY_EVENT":
        oc, ch = OUTKEY.get(int(a["data"]), (f"keycode {a['data']}", "?"))
        return (f"`{ch}`" if ch != "`" else "`` ` ``", a["data"], oc)
    if t == "SYSTEM_ACTION":
        return (SYS.get(a["data"], a["data"]), "—", "—")
    if t == "SHELL_COMMAND":
        desc = ex.get("extra_shell_command_description")
        return (f"Shell: {desc}" if desc else f"Shell: `{decode_shell(a)}`", "—", "—")
    if t == "SOUND":
        return (f"Play sound: {ex.get('extra_sound_file_description', a.get('data', ''))}", "—", "—")
    return (f"{t}: {a.get('data', '')}", "—", "—")


def action_phrase(a):
    """Verb-first plain phrase for the colour-key summary note."""
    t = a["type"]
    ex = extras_of(a)
    if t == "APP":
        return "opens " + APP.get(a["data"], "Open " + a["data"]).replace("Open ", "")
    if t == "SYSTEM_ACTION":
        return {"go_back": "goes back", "go_home": "goes home"}.get(
            a["data"], SYS.get(a["data"], a["data"]))
    if t == "KEY_EVENT":
        _, ch = OUTKEY.get(int(a["data"]), (None, f"keycode {a['data']}"))
        return f"sends {ch}"
    if t == "SOUND":
        return "plays a sound"
    if t == "SHELL_COMMAND":
        sh = decode_shell(a)
        if "service.adb.tcp.port -1" in sh:
            return "disables wireless ADB"
        if "service.adb.tcp.port 5555" in sh:
            return "enables wireless ADB"
        if "service.adb.tcp.port" in sh:
            return "toggles wireless ADB"
        desc = ex.get("extra_shell_command_description")
        return f"runs “{desc}”" if desc else "runs a shell command"
    return t.replace("_", " ").lower()


def colour_note(d):
    """Build the colour-key Notes line from the actual mappings, so it can
    never drift from the table (v1 and v2 map the colours differently)."""
    by = {}  # (keyCode, clickType) -> first matching mapping
    for m in d["keymap_list"]:
        k = m["trigger"]["keys"][0]
        by.setdefault((k["keyCode"], k["clickType"]), m)

    def primary(m):  # the headline action (a Play-sound is only a confirmation)
        real = [a for a in m["actionList"] if a["type"] != "SOUND"] or m["actionList"]
        return action_phrase(real[0])

    parts = []
    for code, name in COLOUR:
        press = by.get((code, 0))
        if not press:
            continue
        s = f"{name} {primary(press)}"
        lp = by.get((code, 1))
        if lp:
            s += f" (long-press: {primary(lp)})"
        parts.append(s)
    return ("**The colour keys are consumed by Key Mapper, not passed to "
            "Harmonium** (they never reach the webview): " + "; ".join(parts) + ".")


def load_data(folder):
    """data.json out of key_mapper.zip (write a readable copy), else an
    existing data.json."""
    zf = folder / "key_mapper.zip"
    if zf.exists():
        with zipfile.ZipFile(zf) as z:
            txt = z.read("data.json").decode("utf-8")
        (folder / "data.json").write_text(txt, encoding="utf-8")
        return json.loads(txt)
    dj = folder / "data.json"
    if dj.exists():
        return json.loads(dj.read_text(encoding="utf-8"))
    return None


def render(d, folder):
    groups = {g["uid"]: g for g in d.get("groups", [])}

    def scope_of(m):
        g = groups.get(m.get("group_uid"))
        if not g:
            return "global"
        cons = []
        for c in g.get("constraints", []):
            if c.get("type") == "constraint_app_foreground":
                pkg = next((x["data"] for x in c.get("extras", [])
                            if x.get("id") == "extra_package_name"), "?")
                cons.append(APP.get(pkg, pkg).replace("Open ", "") + " in foreground")
            else:
                cons.append(c.get("type", "?"))
        return "group “" + g.get("name", "?") + "”" + \
            (" — " + "; ".join(cons) if cons else "")

    rows = []
    shell_cmds = []  # (trigger label, description, decoded command) - for the decoded section
    for m in sorted(d["keymap_list"],
                    key=lambda m: (m["trigger"]["keys"][0]["clickType"],
                                   m["trigger"]["keys"][0]["keyCode"])):
        t = m["trigger"]["keys"][0]
        const, label = KEYCODE.get(t["keyCode"], (f"keycode {t['keyCode']}", f"key {t['keyCode']}"))
        phys = f"{label} {CLICK.get(t['clickType'], t['clickType'])}"
        for a in m["actionList"]:
            if a["type"] == "SHELL_COMMAND":
                shell_cmds.append((phys, extras_of(a).get("extra_shell_command_description", ""),
                                   decode_shell(a)))
        # ONE row per mapping. A multi-action mapping (e.g. Blue = shell + sound)
        # is combined into a single readable cell, not one row per action.
        acts = [describe_action_cell(a) for a in m["actionList"]]
        if len(acts) == 1:
            out, okc, ocon = acts[0]
        else:
            out = " + ".join(a[0] for a in acts)
            okc = ocon = "—"
        rows.append({
            "phys": phys + ("" if m["isEnabled"] else " (disabled)"),
            "ikc": t["keyCode"], "icon": const, "scan": t.get("scanCode", "—"),
            "out": out, "okc": okc, "ocon": ocon, "scope": scope_of(m),
        })

    md = [f"# Astrion Remote Key Map ({folder.name})", "",
          "GENERATED from `key_mapper.zip` — do not hand-edit; rerun",
          "`python ../gen-map-docs.py " + folder.name + "` after mappings change.", "",
          "## KeyMapper rules (what the buttons are remapped to)", "",
          "| Physical key/action | Input keycode | Input Android constant | Scancode | Output/action | Output keycode / package | Output Android constant | Scope |",
          "|---|---:|---|---:|---|---|---|---|"]
    for r in rows:
        md.append(f"| {r['phys']} | {r['ikc']} | `{r['icon']}` | {r['scan']} | "
                  f"{r['out']} | `{r['okc']}` | `{r['ocon']}` | {r['scope']} |"
                  .replace("`—`", "—"))
    md += ["", "## Raw keys (no KeyMapper rule — they reach Harmonium directly)", ""]
    md += ["| Physical key | Emits | What Harmonium does |", "|---|---|---|"]
    for name, emits, does in RAW:
        md.append(f"| {name} | {emits} | {does} |")

    # The exact shell commands, decoded from base64, so a manual-mode setup can
    # retype them (Key Mapper stores them encoded; the table shows only a label).
    if shell_cmds:
        md += ["", "## Shell commands (decoded)", "",
               "The `Shell:` actions in the table run these commands. Key Mapper "
               "stores them base64-encoded; they are decoded here so you can "
               "retype them when building the mappings by hand — in Key Mapper "
               "add a **Shell command** action and paste the line verbatim."]
        for phys, desc, cmd in shell_cmds:
            md += ["", f"### {phys}" + (f" — {desc}" if desc else ""), "",
                   "```sh", cmd, "```"]

    md += ["", "## Notes", "",
           "- `press` is `clickType: 0`; `long-press` is `clickType: 1` — the "
           f"configured long-press delay is {d.get('default_long_press_delay', 600)} ms.",
           "- Scope “global” = the mapping fires everywhere. Grouped mappings "
           "inherit their group's constraints.",
           "- " + colour_note(d),
           "- The hold gestures land on: `]` = hold-Back, `=` = hold-Home, "
           "`F12` = hold-Power (All Off).",
           "- Physical labels derive from the Android `KeyEvent` constants; the "
           "labels printed on the remote may differ.", ""]
    (folder / "astrion-remote-map.md").write_text("\n".join(md), encoding="utf-8")

    n_xlsx = 0
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Astrion"
        ws.append(["Physical key/action", "Input keycode", "Input Android constant",
                   "Scancode", "Output/action", "Output keycode/package",
                   "Output Android constant", "Scope"])
        strip = lambda s: str(s).replace("`", "").strip() or "—"
        for r in rows:
            ws.append([r["phys"], r["ikc"], r["icon"], r["scan"],
                       strip(r["out"]) if r["out"] != "`` ` ``" else "`",
                       r["okc"], r["ocon"], r["scope"]])
        ws2 = wb.create_sheet("Raw keys")
        ws2.append(["Physical key", "Emits", "What Harmonium does"])
        for name, emits, does in RAW:
            ws2.append([name, strip(emits), does])
        if shell_cmds:
            ws3 = wb.create_sheet("Shell commands")
            ws3.append(["Trigger", "Description", "Decoded command"])
            for phys, desc, cmd in shell_cmds:
                ws3.append([phys, desc, cmd])
        wb.save(folder / "KeyCodes Astrion.xlsx")
        n_xlsx = 1
    except ImportError:
        pass
    return len(rows), n_xlsx


def targets():
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        p = Path(arg)
        if not p.is_absolute() and not p.exists():
            p = HERE / arg
        return [p]
    return [p for p in sorted(HERE.glob("v*")) if p.is_dir() and (p / "key_mapper.zip").exists()]


for folder in targets():
    d = load_data(folder)
    if d is None:
        print(f"{folder.name}: no key_mapper.zip / data.json — skipped")
        continue
    nrows, nx = render(d, folder)
    print(f"{folder.name}: {nrows} mapping rows -> astrion-remote-map.md" +
          ("" + " + KeyCodes Astrion.xlsx" if nx else " (xlsx skipped: no openpyxl)"))
